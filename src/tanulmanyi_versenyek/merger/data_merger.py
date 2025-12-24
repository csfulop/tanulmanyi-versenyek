import logging
import json
from pathlib import Path
import pandas as pd

log = logging.getLogger(__name__.split('.')[-1])


def merge_processed_data(cfg):
    """
    Merge all processed CSV files into a single master DataFrame.
    Performs deduplication based on (ev, evfolyam, iskola_nev, helyezes).
    Handles Írásbeli/Szóbeli merge: when both rounds exist for same year+grade,
    drops top N rows from Írásbeli (where N = Szóbeli row count).

    Args:
        cfg: Configuration dictionary

    Returns:
        tuple: (pd.DataFrame, int) - The merged DataFrame and number of duplicates removed
    """
    processed_dir = Path(cfg['paths']['processed_csv_dir'])

    csv_files = list(processed_dir.glob('*.csv'))
    if not csv_files:
        log.warning(f"No CSV files found in {processed_dir}")
        return pd.DataFrame(), 0

    log.info(f"Found {len(csv_files)} CSV files to merge")

    file_data = {}
    for csv_file in csv_files:
        try:
            df = pd.read_csv(csv_file, sep=';', encoding='utf-8')
            filename = csv_file.stem
            parts = filename.split('_')
            year = parts[1]
            grade = parts[2]
            round_type = parts[3]
            key = (year, grade)

            if key not in file_data:
                file_data[key] = {}
            file_data[key][round_type] = df

            log.debug(f"Loaded {csv_file.name}: {len(df)} rows")
        except Exception as e:
            log.error(f"Failed to load {csv_file.name}: {e}")

    if not file_data:
        log.error("No dataframes loaded successfully")
        return pd.DataFrame(), 0

    dataframes = []
    for key, rounds in file_data.items():
        year, grade = key
        if 'szobeli-donto' in rounds and 'irasbeli-donto' in rounds:
            szobeli_df = rounds['szobeli-donto']
            irasbeli_df = rounds['irasbeli-donto']
            szobeli_count = len(szobeli_df)

            irasbeli_filtered = irasbeli_df.iloc[szobeli_count:]
            dataframes.append(szobeli_df)
            dataframes.append(irasbeli_filtered)
            log.debug(f"{year} {grade}: Szóbeli={len(szobeli_df)} rows, Írásbeli={len(irasbeli_df)} rows, dropped top {szobeli_count} from Írásbeli")
        else:
            for round_type, df in rounds.items():
                dataframes.append(df)
                log.debug(f"{year} {grade} {round_type}: {len(df)} rows (no merge needed)")

    master_df = pd.concat(dataframes, ignore_index=True)
    log.info(f"Concatenated all files: {len(master_df)} total rows")

    initial_count = len(master_df)
    master_df = master_df.drop_duplicates(subset=['ev', 'evfolyam', 'iskola_nev', 'helyezes'], keep='first')
    final_count = len(master_df)
    duplicates_removed = initial_count - final_count

    log.info(f"Deduplication complete: removed {duplicates_removed} duplicates, {final_count} rows remaining")

    return master_df, duplicates_removed

def generate_validation_report(df, cfg, duplicates_removed=0, city_stats=None):
    """
    Generate a validation report with data quality metrics.

    Args:
        df: Master DataFrame
        cfg: Configuration dictionary
        duplicates_removed: Number of duplicate rows removed during merge
        city_stats: Optional dictionary with city mapping statistics
    """
    report_path = Path(cfg['paths']['validation_report'])

    total_rows = len(df)
    null_counts = df.isnull().sum().to_dict()
    null_percentages = {col: (count / total_rows * 100) if total_rows > 0 else 0
                       for col, count in null_counts.items()}
    unique_schools = df['iskola_nev'].nunique() if 'iskola_nev' in df.columns else 0

    report = {
        'total_rows': total_rows,
        'duplicates_removed': duplicates_removed,
        'null_counts': null_counts,
        'null_percentages': null_percentages,
        'unique_schools': unique_schools
    }

    if city_stats:
        report['city_mapping'] = {
            'corrections_applied': city_stats.get('corrections_applied', 0),
            'valid_variations': city_stats.get('valid_combinations', 0),
            'unmapped_variations': city_stats.get('unmapped_combinations', 0)
        }

    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)

    log.info(f"Validation report saved to {report_path}")
    log.info(f"Total rows: {total_rows}, Unique schools: {unique_schools}, Duplicates removed: {duplicates_removed}")

def generate_excel_report(df, cfg):
    """
    Generate Excel report with data and summary sheets.
    Creates summary tables instead of pivot tables for better compatibility.

    Args:
        df: Master DataFrame
        cfg: Configuration dictionary
    """
    from pathlib import Path
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    import shutil

    template_path = Path(cfg['paths']['template_file'])
    report_dir = Path(cfg['paths']['report_dir'])
    report_dir.mkdir(parents=True, exist_ok=True)
    output_path = report_dir / 'Bolyai_Analysis_Report.xlsx'

    if not template_path.exists():
        log.error(f"Template file not found: {template_path}")
        return

    shutil.copy(template_path, output_path)
    log.info(f"Copied template to {output_path}")

    wb = load_workbook(output_path)
    ws_data = wb['Data']

    if ws_data.max_row > 1:
        ws_data.delete_rows(2, ws_data.max_row)
        log.debug(f"Cleared existing data rows")

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws_data.cell(row=r_idx, column=c_idx, value=value)

    log.info(f"Wrote {len(df)} rows to Data sheet")

    ws_school = wb.create_sheet("Ranking_by_School")
    school_counts = df.groupby('iskola_nev').size().sort_values(ascending=False)

    ws_school.cell(1, 1, 'iskola_nev').font = Font(bold=True)
    ws_school.cell(1, 2, 'Count').font = Font(bold=True)

    for idx, (school, count) in enumerate(school_counts.items(), start=2):
        ws_school.cell(idx, 1, school)
        ws_school.cell(idx, 2, count)

    ws_school.column_dimensions['A'].width = 60
    ws_school.column_dimensions['B'].width = 10

    log.info(f"Created Ranking_by_School sheet with {len(school_counts)} schools")

    ws_city = wb.create_sheet("Ranking_by_City")
    city_counts = df.groupby('varos').size().sort_values(ascending=False)

    ws_city.cell(1, 1, 'varos').font = Font(bold=True)
    ws_city.cell(1, 2, 'Count').font = Font(bold=True)

    for idx, (city, count) in enumerate(city_counts.items(), start=2):
        ws_city.cell(idx, 1, city)
        ws_city.cell(idx, 2, count)

    ws_city.column_dimensions['A'].width = 40
    ws_city.column_dimensions['B'].width = 10

    log.info(f"Created Ranking_by_City sheet with {len(city_counts)} cities")

    wb.save(output_path)
    log.info(f"Excel report saved to {output_path}")
    wb.close()
